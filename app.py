import bcrypt
import os
import json
from flask import Flask, request, jsonify, session
from flask_cors import CORS
from datetime import date
from zoneinfo import ZoneInfo
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from database import conectar
from database_amt import conectar_amt
from sync import sincronizar_pendentes
from flask import send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from flask_export.export_pdf import backup_pdf as gerar_pdf
from flask_export.export_excel import backup_excel as gerar_excel
from flask_export.export_branding import file_stamp
from flask_export.dados import (
    preparar,
    COLUNAS_CADASTROS,
    COLUNAS_ACESSOS,
)
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
def registrar_log(usuario, acao, descricao, ip):

    try:

        conexao = conectar()
        cursor = conexao.cursor()

        cursor.execute("""
            INSERT INTO logs_auditoria
            (
                usuario,
                acao,
                descricao,
                ip
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s
            )
        """, (
            usuario,
            acao,
            descricao,
            ip
        ))

        conexao.commit()

        cursor.close()
        conexao.close()

    except Exception as erro:

        print("Erro ao registrar log:", erro)

def registrar_backup(tipo, arquivo, destino="LOCAL", status="SUCESSO"):

    try:

        conexao = conectar()
        cursor = conexao.cursor()

        cursor.execute("""
            INSERT INTO backups
            (
                tipo,
                arquivo,
                usuario,
                data_backup,
                destino,
                status
            )
            VALUES
            (
                %s,
                %s,
                %s,
                NOW(),
                %s,
                %s
            )
        """, (
            tipo,
            arquivo,
            session.get("admin"),
            destino,
            status
        ))

        conexao.commit()

        cursor.close()
        conexao.close()

    except Exception as erro:

        print("Erro ao registrar backup:", erro)

app = Flask(__name__)
app.config["SESSION_COOKIE_SAMESITE"] = "None"
app.config["SESSION_COOKIE_SECURE"] = True

limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=[]
)

# ==========================================
# VERIFICAR ADMIN
# ==========================================

def verificar_admin():

    if not session.get("admin"):

        return jsonify({
            "success": False,
            "message": "Não autorizado."
        }), 401

    return None

def criptografar_senha(senha):
    return bcrypt.hashpw(
        senha.encode("utf-8"),
        bcrypt.gensalt()
    ).decode("utf-8")

def verificar_senha(senha, hash_senha):
    return bcrypt.checkpw(
        senha.encode("utf-8"),
        hash_senha.encode("utf-8")
    )
app.secret_key = os.getenv("SECRET_KEY")

SENHA_ADMIN = os.getenv("SENHA_ADMIN", "AMT123456")
CORS(
    app,
    supports_credentials=True,
    origins=["https://portalwifi.lovable.app"]
)


@app.route("/")
def home():
    return "API MEUTRUE funcionando!"


@app.route("/api/cadastro", methods=["POST"])
@limiter.limit("5 per minute")
def cadastro():
    try:
        dados = request.get_json()
        print(dados)

        conexao = conectar()
        cursor = conexao.cursor()

        # Verifica CPF
        cursor.execute(
            "SELECT id FROM usuarios WHERE cpf = %s",
            (dados["cpf"],)
        )

        if cursor.fetchone():
            cursor.close()
            conexao.close()

            return jsonify({
                "sucesso": False,
                "mensagem": "CPF já cadastrado."
            }), 400

        # Verifica e-mail
        cursor.execute(
            "SELECT id FROM usuarios WHERE email = %s",
            (dados["email"],)
        )

        if cursor.fetchone():
            cursor.close()
            conexao.close()

            return jsonify({
                "sucesso": False,
                "mensagem": "E-mail já cadastrado."
            }), 400

        # Data e hora de Fortaleza
        agora = datetime.now(ZoneInfo("America/Fortaleza"))

        latitude, longitude = obter_coordenadas(
             dados.get("rua"),
             dados.get("numero"),
             dados.get("bairro"),
             dados.get("cidade")
        )
        print("===== GEOCODIFICAÇÃO =====")
        print("Número:", dados.get("numero"))
        print("Rua:", dados.get("rua"))
        print("Bairro:", dados.get("bairro"))
        print("Cidade:", dados.get("cidade"))
        print("Latitude:", latitude)
        print("Longitude:", longitude)
        print("==========================")

        sql = """
INSERT INTO usuarios
(
    nome,
    cpf,
    email,
    telefone,
    sexo,
    data_nascimento,
    meio_transporte,
    dias_utilizacao_semana,
    cep,
    rua,
    numero,
    bairro,
    cidade,
    latitude,
    longitude,
    aceite_lgpd,
    data_cadastro
)
VALUES
(
    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
)
"""

        # Campos opcionais
        meio_transporte = dados.get("meio_transporte")

        if meio_transporte:
            meio_transporte = meio_transporte.strip()

        if not meio_transporte:
            meio_transporte = None

        dias_utilizacao = dados.get("dias_utilizacao_semana")

        if dias_utilizacao:
            dias_utilizacao = str(dias_utilizacao).strip()

        if not dias_utilizacao:
            dias_utilizacao = None
        else:
            dias_utilizacao = int(dias_utilizacao)

        valores = (
            dados["nome"],
            dados["cpf"],
            dados["email"],
            dados["telefone"],
            dados["sexo"],
            dados["data_nascimento"],
            meio_transporte,
            dias_utilizacao,
            dados.get("cep") or None,
            dados.get("rua") or None,
            dados.get("numero") or None,
            dados.get("bairro") or None,
            dados.get("cidade") or None,
            latitude,
            longitude,
            1,
            agora
        )

        cursor.execute(sql, valores)

        registro_id = cursor.lastrowid

        conexao.commit()

        adicionar_fila_sincronizacao(registro_id)

        cursor.close()
        conexao.close()
        return jsonify({
            "sucesso": True,
            "mensagem": "Cadastro realizado com sucesso."
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        
        return jsonify({
            "sucesso": False,
            "erro": str(e)
        }), 500

# ============================================================
# FUNÇÃO PARA CONVERTER DATAS EM JSON
# ============================================================

def serialize(valor):
    if isinstance(valor, (datetime, date)):
        return valor.isoformat()
    return valor

# ============================================================
# SINCRONIZAR COM O BANCO DA AMT
# ============================================================

def salvar_no_banco_amt(valores, registro_id):

    try:

        if not os.getenv("AMT_DB_HOST"):
          print("Banco da AMT não configurado.")
          return

        conexao_amt = conectar_amt()
        cursor_amt = conexao_amt.cursor()

        

        sql = """
        INSERT INTO usuarios
        (
            nome,
            cpf,
            email,
            telefone,
            sexo,
            data_nascimento,
            meio_transporte,
            dias_utilizacao_semana,
            cep,
            rua,
            numero,
            bairro,
            cidade,
            latitude,
            longitude,
            aceite_lgpd,
            data_cadastro
        )
        VALUES
        (
            %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s, %s, %s, %s, %s
        )
        """

        cursor_amt.execute(sql, valores)

        conexao_amt.commit()

        cursor_amt.close()
        conexao_amt.close()

        print("Cadastro sincronizado com o banco da AMT.")

    except Exception as erro:

        print("Erro ao sincronizar com a AMT:", erro)

# ============================================================
# ADICIONAR À FILA DE SINCRONIZAÇÃO
# ============================================================

def adicionar_fila_sincronizacao(registro_id):

    try:

        conexao = conectar()
        cursor = conexao.cursor()

        cursor.execute("""
            INSERT INTO fila_sincronizacao
            (
                tabela,
                registro_id
            )
            VALUES
            (
                %s,
                %s
            )
        """, (
            "usuarios",
            registro_id
        ))

        conexao.commit()

        cursor.close()
        conexao.close()

        print("Registro adicionado à fila de sincronização.")

    except Exception as erro:

        print("Erro ao adicionar na fila:", erro)

# ============================================================
# GEOCODIFICAÇÃO
# ============================================================

def obter_coordenadas(rua, numero, bairro, cidade):
    print("FUNÇÃO obter_coordenadas EXECUTOU")

    try:

        geolocator = Nominatim(user_agent="meutrue_maps")
        endereco = f"{rua}, {numero}, {bairro}, {cidade}, Ceará, Brasil"
        print("ENDEREÇO", endereco)

        local = geolocator.geocode(endereco, timeout=10)
        print("RESULTADO", local)

        if local:
            print("LAT", local.latitude)
            print("LON", local.longitude)

            return local.latitude, local.longitude

    except GeocoderTimedOut:
        print("Timeout na geocodificação")

    except Exception as erro:
        print("Erro na geocodificação:", erro)

    return None, None

# ============================================================
# RELATÓRIOS
# ============================================================

@app.route("/api/relatorios", methods=["GET"])
def relatorios():

    resposta = verificar_admin()

    if resposta:
        return resposta

    try:

        conexao = conectar()
        cursor = conexao.cursor()

        # Total
        cursor.execute("SELECT COUNT(*) FROM usuarios")
        total = cursor.fetchone()[0]

        # Hoje
        cursor.execute("""
            SELECT COUNT(*)
            FROM usuarios
            WHERE DATE(data_cadastro)=CURDATE()
        """)
        hoje = cursor.fetchone()[0]

        # Mês
        cursor.execute("""
            SELECT COUNT(*)
            FROM usuarios
            WHERE YEAR(data_cadastro)=YEAR(CURDATE())
            AND MONTH(data_cadastro)=MONTH(CURDATE())
        """)
        mes = cursor.fetchone()[0]

        # Ano
        cursor.execute("""
            SELECT COUNT(*)
            FROM usuarios
            WHERE YEAR(data_cadastro)=YEAR(CURDATE())
        """)
        ano = cursor.fetchone()[0]

        # Cidade
        cursor.execute("""
            SELECT cidade, COUNT(*)
            FROM usuarios
            GROUP BY cidade
            ORDER BY COUNT(*) DESC
        """)

        cidades = []

        for cidade, quantidade in cursor.fetchall():
            cidades.append({
                "cidade": cidade,
                "total": quantidade
            })

        # Bairro
        cursor.execute("""
            SELECT bairro, COUNT(*)
            FROM usuarios
            GROUP BY bairro
            ORDER BY COUNT(*) DESC
        """)

        bairros = []

        for bairro, quantidade in cursor.fetchall():
            bairros.append({
                "bairro": bairro,
                "total": quantidade
            })

        # Cadastros por dia
        cursor.execute("""
            SELECT DATE(data_cadastro), COUNT(*)
            FROM usuarios
            GROUP BY DATE(data_cadastro)
            ORDER BY DATE(data_cadastro)
        """)

        dias = []

        for dia, quantidade in cursor.fetchall():
            dias.append({
                "dia": serialize(dia),
                "total": quantidade
            })

                # Primeiro cadastro
        cursor.execute("SELECT MIN(data_cadastro) FROM usuarios")
        primeiro = serialize(cursor.fetchone()[0])

        # Último cadastro
        cursor.execute("SELECT MAX(data_cadastro) FROM usuarios")
        ultimo = serialize(cursor.fetchone()[0])

        # Quantidade cidades
        cursor.execute("SELECT COUNT(DISTINCT cidade) FROM usuarios")
        total_cidades = cursor.fetchone()[0]

        # Quantidade bairros
        cursor.execute("SELECT COUNT(DISTINCT bairro) FROM usuarios")
        total_bairros = cursor.fetchone()[0]

        # ============================================================
        # Meio de transporte
        # ============================================================

        cursor.execute("""
            SELECT meio_transporte, COUNT(*)
            FROM usuarios
            GROUP BY meio_transporte
            ORDER BY COUNT(*) DESC
        """)

        meios_transporte = []

        for meio, quantidade in cursor.fetchall():
            meios_transporte.append({
                "meio_transporte": meio,
                "total": quantidade
            })

        # ============================================================
        # Dias de utilização por semana
        # ============================================================

        cursor.execute("""
            SELECT dias_utilizacao_semana, COUNT(*)
            FROM usuarios
            GROUP BY dias_utilizacao_semana
            ORDER BY dias_utilizacao_semana
        """)

        dias_utilizacao = []

        for dias_semana, quantidade in cursor.fetchall():
            dias_utilizacao.append({
                "dias": dias_semana,
                "total": quantidade
            })

        cursor.close()
        conexao.close()

        return jsonify({

            "total": total,
            "hoje": hoje,
            "mes": mes,
            "ano": ano,

            "cidades": cidades,

            "bairros": bairros,

            "cadastros_por_dia": dias,

            "primeiro_cadastro": primeiro,

            "ultimo_cadastro": ultimo,

            "total_cidades": total_cidades,

            "total_bairros": total_bairros,

            "meios_transporte": meios_transporte,

            "dias_utilizacao": dias_utilizacao

        })

    except Exception as erro:

        return jsonify({
            "erro": str(erro)
        }), 500
    
# ============================================================
# LISTA DE USUÁRIOS
# ============================================================

@app.route("/api/usuarios", methods=["GET"])
def usuarios():
    resposta = verificar_admin()

    if resposta:
         return resposta

    try:

        conexao = conectar()

        cursor = conexao.cursor(dictionary=True)

        cursor.execute("""

            SELECT *

            FROM usuarios

            ORDER BY data_cadastro DESC

        """)

        dados = cursor.fetchall()

        cursor.close()

        conexao.close()

        return jsonify(dados)

    except Exception as erro:

        return jsonify({

            "erro": str(erro)

        }), 500

# ==========================================
# LOGIN ADMIN
# ==========================================

@app.route("/api/admin/login", methods=["POST"])
@limiter.limit("5 per minute")
def admin_login():
    try:
        dados = request.get_json()

        usuario = dados.get("usuario", "")
        senha = dados.get("password", "")

        conexao = conectar()
        cursor = conexao.cursor(dictionary=True)

        cursor.execute("""
            SELECT *
            FROM administradores
            WHERE usuario=%s
            AND ativo=1
        """, (usuario,))

        admin = cursor.fetchone()

        cursor.close()
        conexao.close()

        if admin and verificar_senha(senha, admin["senha"]):

            session["admin"] = admin["usuario"]

            registrar_log(
                usuario=admin["usuario"],
                acao="LOGIN",
                descricao="Administrador realizou login no sistema.",
                ip=request.remote_addr
            )

            return jsonify({
                "success": True,
                "nome": admin["nome"]
            })

        return jsonify({
            "success": False,
            "message": "Usuário ou senha inválidos."
        }), 401

    except Exception as e:
        return jsonify({
            "success": False,
            "erro": str(e)
        }), 500


# ==========================================
# VERIFICAR LOGIN
# ==========================================

@app.route("/api/admin/verificar", methods=["GET"])
def admin_verificar():

    if session.get("admin"):
        return jsonify({
            "logado": True
        })

    return jsonify({
        "logado": False
    }), 401


# ==========================================
# LOGOUT
# ==========================================

@app.route("/api/admin/logout", methods=["POST"])
def admin_logout():

    usuario = session.get("admin")

    if usuario:
        registrar_log(
            usuario=usuario,
            acao="LOGOUT",
            descricao="Administrador encerrou a sessão.",
            ip=request.remote_addr
        )

    session.clear()

    return jsonify({
        "success": True
    })

# ==========================================
# NOVO ACESSO (USUÁRIO JÁ CADASTRADO)
# ==========================================

@app.route("/api/acesso", methods=["POST"])
@limiter.limit("5 per minute")
def novo_acesso():

    try:

        dados = request.get_json()
        print("Dados recebidos")

        cpf = dados.get("cpf")
        print("CPF:", cpf)

        if not cpf:
            return jsonify({
                "success": False,
                "message": "CPF é obrigatório."
            }), 400

        conexao = conectar()
        cursor = conexao.cursor(dictionary=True)

        # Procura o usuário pelo CPF
        cursor.execute("""
            SELECT *
            FROM usuarios
            WHERE cpf = %s
        """, (cpf,))

        usuario = cursor.fetchone()
        print("USUARIO", usuario)

        if not usuario:
            cursor.close()
            conexao.close()

            return jsonify({
                "success": False,
                "message": "CPF não encontrado."
            }), 404

        # Data e hora atual
        agora = datetime.now(ZoneInfo("America/Fortaleza"))

        # Atualiza o último acesso
        cursor.execute("""
            UPDATE usuarios
            SET ultimo_acesso = %s
            WHERE id = %s
        """, (
            agora,
            usuario["id"]
        ))
# Registra o acesso no histórico
        cursor.execute("""
            INSERT INTO acessos
            (
                usuario_id,
                data_hora,
                onibus,
                linha,
                ip,
                mac_address
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            )
        """, (
            usuario["id"],
            agora,
            None,
            None,
            request.remote_addr,
            None
        ))

        conexao.commit()

        cursor.close()
        conexao.close()

        return jsonify({
          "success": True,
          "nome": usuario["Nome"].split()[0],
          "ultimo_acesso": agora.isoformat()
})

    except Exception as e:
        import traceback

        print("===== ERRO NOVO ACESSO =====")
        traceback.print_exc()
        print("============================")

        return jsonify({
            "success": False,
            "erro": str(e)
        }), 500

# ==========================================
# HISTÓRICO DE ACESSOS
# ==========================================

@app.route("/api/acessos", methods=["GET"])
def listar_acessos():

    resposta = verificar_admin()

    if resposta:
         return resposta

    try:

        conexao = conectar()
        cursor = conexao.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                a.id,
                u.Nome AS nome,
                u.cpf,
                a.data_hora,
                a.ip,
                a.onibus,
                a.linha
            FROM acessos a
            INNER JOIN usuarios u
                ON a.usuario_id = u.id
            ORDER BY a.data_hora DESC
        """)

        acessos = cursor.fetchall()

        cursor.close()
        conexao.close()

        return jsonify(acessos)

    except Exception as erro:

        return jsonify({
            "erro": str(erro)
        }), 500

@app.route("/api/gerar-hash/<senha>")
def gerar_hash(senha):
    return criptografar_senha(senha)


# ============================================================
# EXECUTAR SINCRONIZAÇÃO
# ============================================================

@app.route("/api/sincronizar")
def sincronizar():

    sincronizar_pendentes()

    return jsonify({
        "success": True,
        "mensagem": "Sincronização executada com sucesso."
    })

@app.route("/api/status", methods=["GET"])
def status_sistema():

    try:

        # =====================================
        # Banco principal
        # =====================================

        conexao = conectar()
        cursor = conexao.cursor(dictionary=True)

        banco_online = True

        # =====================================
        # Último backup
        # =====================================

        cursor.execute("""
            SELECT
                tipo,
                arquivo,
                destino,
                status,
                data_backup
            FROM backups
            ORDER BY data_backup DESC
            LIMIT 1
        """)

        ultimo_backup = cursor.fetchone()

        # =====================================
        # Total de backups
        # =====================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM backups
        """)

        total_backups = cursor.fetchone()["total"]

        cursor.close()
        conexao.close()

    except Exception as erro:

        print("Erro status:", erro)

        banco_online = False

        ultimo_backup = None

        total_backups = 0

    # =====================================
    # Banco AMT
    # =====================================

    amt_configurado = all([
        os.getenv("AMT_DB_HOST"),
        os.getenv("AMT_DB_PORT"),
        os.getenv("AMT_DB_USER"),
        os.getenv("AMT_DB_PASSWORD"),
        os.getenv("AMT_DB_NAME")
    ])

    amt_online = False

    if amt_configurado:

        try:

            conexao_amt = conectar_amt()

            conexao_amt.close()

            amt_online = True

        except Exception:

            amt_online = False

    # =====================================
    # Resposta
    # =====================================

    return jsonify({

        "api": {
            "online": True,
            "latencia": 180
        },

        "banco_principal": {
            "online": banco_online,
            "conexoes": "Saudáveis" if banco_online else "Falha na conexão"
        },

        "banco_amt": {
            "configurado": amt_configurado,
            "online": amt_online
        },

        "sincronizador": {
            "online": True,
            "ultima_execucao": None
        },

        "backup": {

            "ultimo": (
                ultimo_backup["data_backup"].strftime("%d/%m/%Y %H:%M")
                if ultimo_backup else None
            ),

            "tipo": (
                ultimo_backup["tipo"]
                if ultimo_backup else None
            ),

            "arquivo": (
                ultimo_backup["arquivo"]
                if ultimo_backup else None
            ),

            "destino": (
                ultimo_backup["destino"]
                if ultimo_backup else "LOCAL"
            ),

            "status": (
                ultimo_backup["status"]
                if ultimo_backup else "AGUARDANDO"
            ),

            "total": total_backups

        }

    }), 200

@app.route("/api/sincronizacao/status", methods=["GET"])
def status_sincronizacao():

    try:

        conexao = conectar()
        cursor = conexao.cursor(dictionary=True)

        # Pendentes
        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM fila_sincronizacao
            WHERE status = 'PENDENTE'
        """)
        pendentes = cursor.fetchone()["total"]

        # Sincronizados
        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM fila_sincronizacao
            WHERE status = 'SINCRONIZADO'
        """)
        sincronizados = cursor.fetchone()["total"]

        # Erros
        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM fila_sincronizacao
            WHERE status = 'ERRO'
        """)
        erros = cursor.fetchone()["total"]

        # Última sincronização
        cursor.execute("""
            SELECT MAX(data_sincronizacao) AS ultima
            FROM fila_sincronizacao
        """)

        ultima = cursor.fetchone()["ultima"]

        cursor.close()
        conexao.close()

        return jsonify({
            "configurado": all([
                os.getenv("AMT_DB_HOST"),
                os.getenv("AMT_DB_PORT"),
                os.getenv("AMT_DB_USER"),
                os.getenv("AMT_DB_PASSWORD"),
                os.getenv("AMT_DB_NAME")
            ]),
            "pendentes": pendentes,
            "sincronizados": sincronizados,
            "erros": erros,
            "ultima_sincronizacao": ultima
        })

    except Exception as erro:

        print("ERRO SINCRONIZAÇÃO:", erro)

        return jsonify({
            "configurado": False,
            "pendentes": 0,
            "sincronizados": 0,
            "erros": 0,
            "ultima_sincronizacao": None
        }), 500

@app.route("/api/logs", methods=["POST"])
def criar_log():

    try:

        if not session.get("admin"):
            return jsonify({
                "success": False,
                "message": "Não autorizado."
            }), 401

        dados = request.get_json()

        registrar_log(
            usuario=session.get("admin"),
            acao=dados.get("acao"),
            descricao=dados.get("descricao"),
            ip=request.remote_addr
        )

        return jsonify({
            "success": True
        })

    except Exception as erro:

        return jsonify({
            "success": False,
            "erro": str(erro)
        }), 500

@app.route("/api/backup", methods=["POST"])
def gerar_backup():

    resposta = verificar_admin()
    if resposta:
        return resposta

    try:

        conexao = conectar()
        cursor = conexao.cursor(dictionary=True)

        backup = {
            "sistema": "Portal WiFi TRUE",
            "versao": "1.0",
            "gerado_por": session.get("admin"),
            "data_backup": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "tabelas": {}
        }

        # Descobre todas as tabelas do banco
        cursor.execute("SHOW TABLES")

        tabelas = []

        for linha in cursor.fetchall():
            nome_tabela = list(linha.values())[0]
            tabelas.append(nome_tabela)

        # Faz backup de cada tabela
        for tabela in tabelas:

            cursor.execute(f"SELECT * FROM `{tabela}`")

            registros = cursor.fetchall()

            # Converte datas para texto
            dados = []

            for registro in registros:

                novo = {}

                for chave, valor in registro.items():

                    if isinstance(valor, (datetime, date)):
                        novo[chave] = valor.isoformat()

                    else:
                        novo[chave] = valor

                dados.append(novo)

            backup["tabelas"][tabela] = dados

        cursor.close()
        conexao.close()

        registrar_log(
            usuario=session.get("admin"),
            acao="GERAR_BACKUP",
            descricao="Administrador gerou backup completo do banco.",
            ip=request.remote_addr
        )
        registrar_backup(
            tipo="JSON",
            arquivo=nome,
            destino="LOCAL",
            status="SUCESSO"
        )

        json_backup = json.dumps(
            backup,
            ensure_ascii=False,
            indent=4
        )

        arquivo = BytesIO(
            json_backup.encode("utf-8")
        )

        nome = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

        return send_file(
            arquivo,
            as_attachment=True,
            download_name=nome,
            mimetype="application/json"
        )

    except Exception as erro:

        return jsonify({
            "success": False,
            "erro": str(erro)
        }), 500

# ==========================================
# ESTILOS DO EXCEL
# ==========================================

header_fill = PatternFill(
    fill_type="solid",
    fgColor="03D199"
)

header_font = Font(
    bold=True,
    color="FFFFFF",
    size=11
)

titulo_font = Font(
    bold=True,
    size=18,
    color="FFFFFF"
)

subtitulo_font = Font(
    italic=True,
    size=10,
    color="64748B"
)

texto_font = Font(
    size=10,
    color="0B1F1A"
)

kpi_font = Font(
    bold=True,
    size=12,
    color="0A9B73"
)

borda = Border(

    bottom=Side(
        style="thin",
        color="E2E8F0"
    )
)

centro = Alignment(
    horizontal="center",
    vertical="center"
)

# ==========================================
# ESTILIZA UMA PLANILHA
# ==========================================

def estilizar_planilha(ws):

    # Congela a primeira linha
    ws.freeze_panes = "A2"

    # Ativa filtros
    ws.auto_filter.ref = ws.dimensions

    # Cabeçalho
    for celula in ws[1]:

        celula.fill = header_fill
        celula.font = header_font
        celula.border = borda
        celula.alignment = centro

    # Zebra nas linhas
    for linha in range(2, ws.max_row + 1):

        if linha % 2 == 0:

            for celula in ws[linha]:

                celula.fill = PatternFill(
                    fill_type="solid",
                    fgColor="F8FAFC"
                )

    # Centraliza o ID
    for linha in range(2, ws.max_row + 1):

        ws[f"A{linha}"].alignment = centro

def carregar_dados_exportacao():
    """
    Carrega os dados necessários para as exportações.
    Reutiliza a conexão existente do sistema.
    """

    conexao = conectar()
    cursor = conexao.cursor(dictionary=True)

    # ============================
    # USUÁRIOS
    # ============================

    cursor.execute("""
        SELECT *
        FROM usuarios
        ORDER BY data_cadastro DESC
    """)

    usuarios = cursor.fetchall()

    # ============================
    # ACESSOS
    # ============================

    cursor.execute("""
        SELECT
            a.id,
            u.nome,
            u.cpf,
            a.ip,
            a.data_hora AS data_acesso,
            a.onibus,
            a.linha
        FROM acessos a
        INNER JOIN usuarios u
            ON u.id = a.usuario_id
        ORDER BY a.data_hora DESC
    """)

    acessos = cursor.fetchall()

    cursor.close()
    conexao.close()

    return usuarios, acessos

@app.route("/api/backup/excel", methods=["POST"])
def backup_excel():

    resposta = verificar_admin()

    if resposta:
        return resposta

    try:

        usuarios, acessos = carregar_dados_exportacao()

        cadastros, acessos, kpis, graficos = preparar(
            usuarios,
            acessos
        )

        meta = {
            "titulo": "Relatório Executivo",
            "usuario": session.get("admin")
        }

        abas = [

            {
                "nome": "Acessos",
                "colunas": COLUNAS_ACESSOS,
                "linhas": acessos
            }

        ]

        arquivo = gerar_excel(

            meta=meta,

            kpis=kpis,

            colunas=COLUNAS_CADASTROS,

            linhas=cadastros,

            sheet_name="Cadastros",

            graficos=graficos,

            abas=abas

        )

        registrar_log(

            usuario=session.get("admin"),

            acao="BACKUP_EXCEL",

            descricao="Administrador exportou relatório Excel.",

            ip=request.remote_addr

        )
        registrar_backup(
            tipo="EXCEL",

            arquivo=f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",

            destino="LOCAL",

            status="SUCESSO"
        )

        nome = f"portal_true_{file_stamp()}.xlsx"

        return send_file(

            arquivo,

            as_attachment=True,

            download_name=nome,

            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        )
 
    except Exception as erro:

        return jsonify({

            "success": False,

            "erro": str(erro)

        }), 500

@app.route("/api/backup/pdf", methods=["POST"])
def backup_pdf():

    resposta = verificar_admin()

    if resposta:
        return resposta

    try:

        usuarios, acessos = carregar_dados_exportacao()

        cadastros, acessos, kpis, graficos = preparar(
            usuarios,
            acessos
        )

        meta = {
            "titulo": "Relatório Executivo",
            "usuario": session.get("admin")
        }

        head = [
            coluna["header"]
            for coluna in COLUNAS_CADASTROS
        ]

        body = [

            [

                registro.get(coluna["key"], "")

                for coluna in COLUNAS_CADASTROS

            ]

            for registro in cadastros

        ]

        secoes = []

        if acessos:

            secoes.append({

                "titulo": "Histórico de Acessos",

                "head": [

                    coluna["header"]

                    for coluna in COLUNAS_ACESSOS

                ],

                "body": [

                    [

                        registro.get(coluna["key"], "")

                        for coluna in COLUNAS_ACESSOS

                    ]

                    for registro in acessos

                ]

            })

        arquivo = gerar_pdf(

            meta,

            kpis,

            head,

            body,

            graficos,

            secoes

        )

        registrar_log(

            session.get("admin"),

            "EXPORTAR_PDF",

            "Relatório PDF exportado.",

            request.remote_addr

        )
        registrar_backup(
            tipo="PDF",
            arquivo=f"portal-true-{file_stamp()}.pdf",
            destino="LOCAL",
            status="SUCESSO"
        )

        return send_file(

            arquivo,

            as_attachment=True,

            download_name=f"portal-true-{file_stamp()}.pdf",

            mimetype="application/pdf"

        )

    except Exception as erro:

        return jsonify({

            "success": False,

            "erro": str(erro)

        }), 500

if __name__ == "__main__":
    print("API iniciando...")
    app.run(host="0.0.0.0", port=5000, debug=True)