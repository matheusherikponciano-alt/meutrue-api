# -*- coding: utf-8 -*-

"""
Geração do Dashboard Executivo em Excel (openpyxl).
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.properties import PageSetupProperties

from .export_branding import (
    BRAND,
    ORG,
    FONT_XLSX,
    PALETA,
    argb,
    rgb_hex,
    logo_path,
    now_fortaleza,
)

COLS = 8


# ==========================================================
# ESTILOS
# ==========================================================

def _fill(cor):
    return PatternFill(
        fill_type="solid",
        start_color=argb(cor),
        end_color=argb(cor),
    )


def _font(
    tamanho=10,
    bold=False,
    italic=False,
    color=BRAND["ink"],
):
    return Font(
        name=FONT_XLSX,
        size=tamanho,
        bold=bold,
        italic=italic,
        color=rgb_hex(color),
    )


def _align(
    horizontal="left",
    indent=1,
):
    return Alignment(
        horizontal=horizontal,
        vertical="center",
        indent=indent,
    )


def _side(style="thin", color=BRAND["border"]):
    return Side(
        style=style,
        color=rgb_hex(color),
    )


def _mesclar(ws, row, col1, col2):

    if col2 > col1:

        ws.merge_cells(
            start_row=row,
            start_column=col1,
            end_row=row,
            end_column=col2,
        )

    return ws.cell(
        row=row,
        column=col1,
    )


def _pintar(ws, row, total_colunas, cor):

    for coluna in range(1, total_colunas + 1):

        ws.cell(
            row=row,
            column=coluna,
        ).fill = _fill(cor)


def _rodape(ws):

    ws.oddFooter.left.text = (
        f"{ORG['system']} · AMT Eusébio"
    )

    ws.oddFooter.center.text = "&D"

    ws.oddFooter.right.text = (
        "Página &P de &N"
    )


def _paisagem(ws):

    ws.page_setup.orientation = "landscape"

    ws.page_setup.fitToWidth = 1

    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr = (
        PageSetupProperties(
            fitToPage=True
        )
    )


# ==========================================================
# DASHBOARD EXECUTIVO
# ==========================================================

def _dashboard(
    wb,
    meta,
    kpis,
    graficos,
):

    ws = wb.create_sheet(
        title="Dashboard Executivo"
    )

    ws.sheet_view.showGridLines = False

    _paisagem(ws)

    for coluna in range(1, COLS + 1):

        ws.column_dimensions[
            get_column_letter(coluna)
        ].width = 20

    caminho_logo = logo_path()

    if caminho_logo:

        try:

            from openpyxl.drawing.image import Image

            logo = Image(caminho_logo)

            logo.width = 95

            logo.height = 95

            ws.add_image(
                logo,
                "A1"
            )

        except Exception:

            pass

    def faixa(
        linha,
        texto,
        tamanho,
        bold=False,
        cor=BRAND["white"],
    ):

        cel = _mesclar(
            ws,
            linha,
            1,
            COLS,
        )

        cel.value = texto

        cel.font = _font(
            tamanho=tamanho,
            bold=bold,
            color=cor,
        )

        cel.alignment = _align(indent=9)

        _pintar(
            ws,
            linha,
            COLS,
            BRAND["ink"],
        )

    faixa(
        1,
        "PORTAL WI-FI TRUE",
        20,
        True,
    )

    faixa(
        2,
        ORG["org_display"],
        10,
    )

    faixa(
        3,
        "Relatório Executivo",
        11,
        True,
        BRAND["yellow"],
    )

    faixa(
        4,
        "Gerado em " + now_fortaleza(),
        9,
    )

    faixa(
        5,
        "Exportado por "
        + str(meta.get("usuario")),
        9,
    )

    _pintar(
        ws,
        6,
        COLS,
        BRAND["green"],
    )

    linha = 9

    titulo = _mesclar(
        ws,
        linha,
        1,
        COLS,
    )

    titulo.value = "INDICADORES"

    titulo.font = _font(
        10,
        True,
        color=BRAND["greenDark"],
    )

    linha += 1

    por_linha = 4

    # ==========================================================
    # CARDS DE KPI
    # ==========================================================

    for i, kpi in enumerate(kpis):

        linha_card = linha + (i // por_linha) * 3

        coluna = (i % por_linha) * 2 + 1

        label = _mesclar(
            ws,
            linha_card,
            coluna,
            coluna + 1,
        )

        label.value = str(
            kpi["label"]
        ).upper()

        label.font = _font(
            8,
            True,
            color=BRAND["slate"],
        )

        label.alignment = _align()

        valor = _mesclar(
            ws,
            linha_card + 1,
            coluna,
            coluna + 1,
        )

        valor.value = kpi["value"]

        valor.font = _font(
            17,
            True,
        )

        valor.alignment = _align()

        for c in (coluna, coluna + 1):

            ws.cell(
                row=linha_card,
                column=c,
            ).fill = _fill(BRAND["greenSoft"])

            ws.cell(
                row=linha_card + 1,
                column=c,
            ).fill = _fill(BRAND["greenSoft"])

    if kpis:

        linha += ((len(kpis) + 3) // 4) * 3 + 2

    # ==========================================================
    # GRÁFICOS
    # ==========================================================

    cab = _mesclar(
        ws,
        linha,
        1,
        COLS,
    )

    cab.value = "ANÁLISES"

    cab.font = _font(
        10,
        True,
        color=BRAND["greenDark"],
    )

    linha += 2

    for grafico in graficos or []:

        itens = sorted(

            [
                x for x in grafico["dados"]

                if x["value"] > 0
            ],

            key=lambda x: x["value"],

            reverse=True

        )[:10]

        titulo = _mesclar(
            ws,
            linha,
            1,
            COLS,
        )

        titulo.value = grafico["titulo"]

        titulo.font = _font(
            11,
            True,
            color=BRAND["white"],
        )

        titulo.alignment = _align()

        _pintar(
            ws,
            linha,
            COLS,
            BRAND["greenDark"],
        )

        linha += 1

        if len(itens) < grafico.get(
            "minimo",
            2,
        ):

            aviso = _mesclar(
                ws,
                linha,
                1,
                COLS,
            )

            aviso.value = (
                "Não existem dados suficientes para gerar esta análise."
            )

            aviso.font = _font(
                10,
                italic=True,
                color=BRAND["slate"],
            )

            aviso.fill = _fill(
                BRAND["surface"]
            )

            linha += 2

            continue

        cabecalhos = [

            "Categoria",

            "Quantidade",

            "%",

            "Distribuição"

        ]

        for c, texto in enumerate(
            cabecalhos,
            start=1,
        ):

            cel = ws.cell(
                row=linha,
                column=c,
            )

            cel.value = texto

            cel.font = _font(
                9,
                True,
            )

            cel.fill = _fill(
                BRAND["greenSoft"]
            )

        _mesclar(
            ws,
            linha,
            4,
            COLS,
        )

        linha += 1

        total = sum(
            x["value"]
            for x in itens
        )

        maior = max(
            x["value"]
            for x in itens
        )

        inicio_barra = linha

        for indice, item in enumerate(itens):

            ws.cell(
                row=linha,
                column=1,
            ).value = item["name"]

            ws.cell(
                row=linha,
                column=2,
            ).value = item["value"]

            ws.cell(
                row=linha,
                column=3,
            ).value = item["value"] / total

            ws.cell(
                row=linha,
                column=3,
            ).number_format = "0.0%"

            barra = "█" * max(
                1,
                int(
                    (item["value"] / maior) * 28
                )
            )

            cel = ws.cell(
                row=linha,
                column=4,
            )

            cel.value = barra

            cel.font = _font(
                color=PALETA[
                    indice % len(PALETA)
                ]
            )

            _mesclar(
                ws,
                linha,
                4,
                COLS,
            )

            if linha % 2 == 0:

                for c in range(
                    1,
                    COLS + 1,
                ):

                    ws.cell(
                        row=linha,
                        column=c,
                    ).fill = _fill(
                        BRAND["surface"]
                    )

            linha += 1

        ws.conditional_formatting.add(

            f"B{inicio_barra}:B{linha-1}",

            DataBarRule(

                start_type="min",

                end_type="max",

                color=rgb_hex(
                    BRAND["green"]
                )

            )

        )

        linha += 2
        
    # ==========================================================
    # RODAPÉ
    # ==========================================================

    rodape = _mesclar(
        ws,
        linha,
        1,
        COLS,
    )

    rodape.value = ORG["footer"]

    rodape.font = _font(
        9,
        italic=True,
        color=BRAND["slate"],
    )

    rodape.alignment = _align()

    _rodape(ws)

    return ws


# ==========================================================
# PLANILHAS DE DADOS
# ==========================================================

def _planilha_dados(
    wb,
    nome,
    colunas,
    linhas,
):

    ws = wb.create_sheet(
        title=nome[:31]
    )

    ws.sheet_view.showGridLines = False

    _paisagem(ws)

    for i, coluna in enumerate(colunas, start=1):

        letra = get_column_letter(i)

        ws.column_dimensions[
            letra
        ].width = coluna.get(
            "width",
            20,
        )

        cab = ws.cell(
            row=1,
            column=i,
        )

        cab.value = coluna["header"]

        cab.font = _font(
            10,
            True,
            color=BRAND["white"],
        )

        cab.fill = _fill(
            BRAND["greenDark"]
        )

        cab.alignment = _align()

    linha_excel = 2

    for registro in linhas:

        for coluna, info in enumerate(
            colunas,
            start=1,
        ):

            valor = registro.get(
                info["key"],
                "",
            )

            cel = ws.cell(
                row=linha_excel,
                column=coluna,
            )

            cel.value = valor

            cel.font = _font()

            if linha_excel % 2 == 0:

                cel.fill = _fill(
                    BRAND["surface"]
                )

        linha_excel += 1

    ws.freeze_panes = "A2"

    if linha_excel > 2:

        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(colunas))}{linha_excel-1}"
        )

    _rodape(ws)

    return ws


# ==========================================================
# EXPORTAÇÃO
# ==========================================================

def backup_excel(
    meta,
    kpis,
    colunas,
    linhas,
    sheet_name,
    graficos=None,
    abas=None,
):

    wb = Workbook()

    wb.remove(
        wb.active
    )

    wb.properties.creator = ORG["system"]

    wb.properties.company = ORG["org_display"]

    wb.properties.title = meta.get(
        "titulo",
        "Relatório Executivo",
    )

    _dashboard(
        wb,
        meta,
        kpis,
        graficos,
    )

    _planilha_dados(
        wb,
        sheet_name,
        colunas,
        linhas,
    )

    for aba in abas or []:

        _planilha_dados(

            wb,

            aba["nome"],

            aba["colunas"],

            aba["linhas"]

        )

    arquivo = BytesIO()

    wb.save(
        arquivo
    )

    arquivo.seek(0)

    return arquivo